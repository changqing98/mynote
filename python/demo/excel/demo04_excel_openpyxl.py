import openpyxl



def main():
    data = openpyxl.Workbook()  # 新建工作簿
    data.create_sheet('Sheet1')  # 添加页
    # table = data.get_sheet_by_name('Sheet1') # 获得指定名称页
    table = data.active  # 获得当前活跃的工作页，默认为第一个工作页
    # table.cell("client_id", "code", 'crm_id', "activation_count", "license_count")  # 行，列，值 这里是从1开始计数的
    table.cell(1, 1, "client_id")
    table.cell(1, 2, "code")
    table.cell(1, 3, "crm_id")
    table.cell(1, 4, "activation_count")
    table.cell(1, 5, "license_count")
    file1 = open("./resources/clients.txt")
    file2 = open("./resources/clients2.txt")

    lines1 = file1.readlines()

    lines2 = file2.readlines()

    col = 2
    for i in range(3, len(lines1) - 1):
        str1 = lines1[i].split("|")
        str2 = lines2[i].split("|")

        id1 = str1[1]
        code = str1[2]
        crm_id = str1[3]
        id2 = str2[1]
        activation_count = str2[2]
        license_count = str2[3]

        table.cell(col, 1, int(id1))
        table.cell(col, 2, code)
        table.cell(col, 3, int(crm_id))
        table.cell(col, 4, int(activation_count))
        table.cell(col, 5, int(license_count))

        col = col+1

        print(id1, id2, code, crm_id, activation_count, license_count)
        print("\n")
    data.save("./client.xlsx")


if __name__ == '__main__':
    main()
